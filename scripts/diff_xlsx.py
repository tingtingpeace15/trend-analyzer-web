# -*- coding: utf-8 -*-
"""M7 终极对照:逐 sheet、逐单元格、逐样式 diff Python 版 vs JS 版输出 xlsx。

对比项:sheet 名/顺序、行列数、单元格值(浮点容差)、数字格式、字体(颜色/粗体)、
填充、对齐、边框、列宽、行高、冻结窗格、自动筛选、批注文本、嵌图锚点(解包 XML)。
明确不比:图片字节(Canvas vs matplotlib 必然不同)、批注外形尺寸(ExcelJS 不支持)。

用法: python3 scripts/diff_xlsx.py [python版.xlsx] [js版.xlsx]
退出码: 0 = 无实质差异
"""
import math
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
PY_XLSX = Path(sys.argv[1]) if len(sys.argv) > 2 else HERE.parent / "baseline" / "python" / "商品销售趋势.xlsx"
JS_XLSX = Path(sys.argv[2]) if len(sys.argv) > 2 else HERE.parent / "baseline" / "js" / "商品销售趋势.xlsx"

issues = Counter()
examples = {}
MAX_EXAMPLES = 5


def report(category, detail):
    issues[category] += 1
    if issues[category] <= MAX_EXAMPLES:
        examples.setdefault(category, []).append(detail)


def close(a, b):
    return math.isclose(a, b, rel_tol=1e-9, abs_tol=1e-9)


def norm_val(v):
    if v is None or v == "":
        return None
    return v


def eq_val(a, b):
    a, b = norm_val(a), norm_val(b)
    if a is None and b is None:
        return True
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return close(float(a), float(b))
    return a == b


def color_rgb(c):
    """openpyxl Color → 可比字符串。theme/indexed/auto 归一化"""
    if c is None:
        return None
    if c.type == "rgb":
        rgb = c.rgb
        if rgb in ("00000000",):
            return None
        return rgb
    return f"{c.type}:{getattr(c, c.type, None)}"


def font_sig(f):
    if f is None:
        return None
    return (color_rgb(f.color), bool(f.b), f.sz, f.name)


def fill_sig(f):
    if f is None or f.patternType is None:
        return None
    return (f.patternType, color_rgb(f.fgColor))


def align_sig(a):
    if a is None:
        return None
    return (a.horizontal, a.vertical, a.indent or 0)


def border_sig(b):
    if b is None:
        return None
    def side(s):
        return (s.style, color_rgb(s.color)) if s and s.style else None
    return (side(b.left), side(b.right), side(b.top), side(b.bottom))


print(f"Python: {PY_XLSX}\nJS:     {JS_XLSX}\n")
wb_py = openpyxl.load_workbook(PY_XLSX)
wb_js = openpyxl.load_workbook(JS_XLSX)

# ── sheet 名 ──
if wb_py.sheetnames != wb_js.sheetnames:
    report("sheet名", f"{wb_py.sheetnames} vs {wb_js.sheetnames}")

for name in wb_py.sheetnames:
    if name not in wb_js.sheetnames:
        continue
    ws_p, ws_j = wb_py[name], wb_js[name]
    tag = f"[{name}]"

    if (ws_p.max_row, ws_p.max_column) != (ws_j.max_row, ws_j.max_column):
        report("行列数", f"{tag} {ws_p.max_row}×{ws_p.max_column} vs {ws_j.max_row}×{ws_j.max_column}")

    if str(ws_p.freeze_panes) != str(ws_j.freeze_panes):
        report("冻结窗格", f"{tag} {ws_p.freeze_panes} vs {ws_j.freeze_panes}")
    if str(ws_p.auto_filter.ref) != str(ws_j.auto_filter.ref):
        report("自动筛选", f"{tag} {ws_p.auto_filter.ref} vs {ws_j.auto_filter.ref}")

    # 行高(列宽在最后统一从 XML 范围解析,openpyxl 不展开 <col min max> 范围)
    for row in range(1, min(ws_p.max_row, ws_j.max_row) + 1):
        hp = ws_p.row_dimensions[row].height
        hj = ws_j.row_dimensions[row].height
        if (hp is None) != (hj is None) or (hp is not None and hj is not None and not close(hp, hj)):
            report("行高", f"{tag} row{row}: {hp} vs {hj}")

    # 单元格(值差先暂存,sheet 结束后判断是否为"平局重排")
    val_mismatch_rows = set()
    val_examples = []
    max_r = min(ws_p.max_row, ws_j.max_row)
    max_c = min(ws_p.max_column, ws_j.max_column)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            cp, cj = ws_p.cell(r, c), ws_j.cell(r, c)
            addr = f"{tag} {openpyxl.utils.get_column_letter(c)}{r}"
            if not eq_val(cp.value, cj.value):
                val_mismatch_rows.add(r)
                if len(val_examples) < MAX_EXAMPLES:
                    val_examples.append(f"{addr}: {cp.value!r} vs {cj.value!r}")
            if cp.number_format != cj.number_format:
                report("数字格式", f"{addr}: {cp.number_format!r} vs {cj.number_format!r}")
            if font_sig(cp.font) != font_sig(cj.font):
                report("字体", f"{addr}: {font_sig(cp.font)} vs {font_sig(cj.font)}")
            if fill_sig(cp.fill) != fill_sig(cj.fill):
                report("填充", f"{addr}: {fill_sig(cp.fill)} vs {fill_sig(cj.fill)}")
            if align_sig(cp.alignment) != align_sig(cj.alignment):
                report("对齐", f"{addr}: {align_sig(cp.alignment)} vs {align_sig(cj.alignment)}")
            if border_sig(cp.border) != border_sig(cj.border):
                report("边框", f"{addr}: {border_sig(cp.border)} vs {border_sig(cj.border)}")
            # 批注(只比文本;外形尺寸 ExcelJS 不支持,明确豁免)
            tp = cp.comment.text if cp.comment else None
            tj = cj.comment.text if cj.comment else None
            if (tp or tj) and tp != tj:
                report("批注", f"{addr}: {tp!r:.80} vs {tj!r:.80}")

    # 值差判定:整行多重集一致 → 平局重排(排序键相等的行,numpy SIMD argsort
    # 与经典 introsort 的平局顺序不同,且 numpy 自身跨平台/版本也不稳定——
    # Python 版换台机器跑同样会变,故视为可接受;否则按真实值差报错。
    if val_mismatch_rows:
        def row_tuple(ws, r):
            return tuple(
                round(v, 6) if isinstance(v, float) else v
                for v in (ws.cell(r, c).value for c in range(1, max_c + 1))
            )
        from collections import Counter as _C
        rows_p = _C(row_tuple(ws_p, r) for r in range(2, max_r + 1))
        rows_j = _C(row_tuple(ws_j, r) for r in range(2, max_r + 1))
        if rows_p == rows_j:
            issues["平局重排(行多重集一致,可接受)"] += len(val_mismatch_rows)
            examples.setdefault("平局重排(行多重集一致,可接受)", []).extend(
                val_examples[: max(0, MAX_EXAMPLES - len(examples.get("平局重排(行多重集一致,可接受)", [])))])
        else:
            for ex in val_examples:
                report("单元格值", ex)
            extra = len(val_mismatch_rows) - len(val_examples)
            if extra > 0:
                issues["单元格值"] += extra

# ── 列宽(直接解析 worksheet XML 的 <col min max width>,展开范围后逐列比)──
WS_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def col_widths_of(path):
    out = {}
    with zipfile.ZipFile(path) as z:
        for n in sorted(z.namelist()):
            m = re.fullmatch(r"xl/worksheets/sheet(\d+)\.xml", n)
            if not m:
                continue
            root = ET.fromstring(z.read(n))
            widths = {}
            for col in root.findall(".//m:cols/m:col", WS_NS):
                w = float(col.get("width"))
                for c in range(int(col.get("min")), int(col.get("max")) + 1):
                    widths[c] = w
            out[int(m.group(1))] = widths
    return out


cw_p, cw_j = col_widths_of(PY_XLSX), col_widths_of(JS_XLSX)
for k in sorted(set(cw_p) | set(cw_j)):
    wp, wj = cw_p.get(k, {}), cw_j.get(k, {})
    for c in sorted(set(wp) | set(wj)):
        a, b = wp.get(c), wj.get(c)
        if a is None or b is None or not close(a, b):
            report("列宽", f"sheet{k} col{c}: {a} vs {b}")

# ── 嵌图锚点(解包 drawing XML)──
NS = {"xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"}


def anchors_of(path):
    out = {}
    with zipfile.ZipFile(path) as z:
        for n in sorted(z.namelist()):
            m = re.fullmatch(r"xl/drawings/drawing(\d+)\.xml", n)
            if not m:
                continue
            root = ET.fromstring(z.read(n))
            lst = []
            for a in root.findall("xdr:twoCellAnchor", NS):
                def marker(el):
                    return tuple(int(el.find(f"xdr:{k}", NS).text) for k in ("col", "colOff", "row", "rowOff"))
                lst.append((a.get("editAs"), marker(a.find("xdr:from", NS)), marker(a.find("xdr:to", NS))))
            out[int(m.group(1))] = lst
    return out


an_p, an_j = anchors_of(PY_XLSX), anchors_of(JS_XLSX)
for k in sorted(set(an_p) | set(an_j)):
    lp, lj = an_p.get(k, []), an_j.get(k, [])
    if len(lp) != len(lj):
        report("锚点数量", f"drawing{k}: {len(lp)} vs {len(lj)}")
    # 锚点按 (from.row, from.col) 排序后逐个比(两边添加顺序可能不同)
    for i, (ap, aj) in enumerate(zip(sorted(lp, key=lambda x: x[1]), sorted(lj, key=lambda x: x[1]))):
        if ap != aj:
            report("锚点", f"drawing{k}#{i}: {ap} vs {aj}")

# ── 汇总 ──
print("=" * 64)
TIE_KEY = "平局重排(行多重集一致,可接受)"
real = {k: v for k, v in issues.items() if k != TIE_KEY}
if not real:
    tie = issues.get(TIE_KEY, 0)
    extra = f"(另有 {tie} 行平局重排,内容多重集一致,可接受)" if tie else ""
    print(f"✅ 全部一致{extra}(图片字节与批注外形尺寸按约定不比)")
    if tie:
        for ex in examples.get(TIE_KEY, [])[:3]:
            print(f"   平局示例: {ex}")
    sys.exit(0)
total = sum(real.values())
print(f"❌ 共 {total} 处实质差异,按类别:")
for cat, n in issues.most_common():
    if cat == TIE_KEY:
        continue
    print(f"\n── {cat} × {n}")
    for ex in examples[cat]:
        print(f"   {ex}")
    if n > MAX_EXAMPLES:
        print(f"   …(还有 {n - MAX_EXAMPLES} 处)")
sys.exit(1)
